# this is for adding the "external" folder to the system path, because the QGIS python is only looking in the system path for python packages
# you can see that this works by doing import sys, sys.path inside the QGIS python console
import os, sys
sys.path.append(os.path.join(os.path.dirname(__file__), "external"))

import openpyxl
#import win32com.client
import io
import xlwings as xw


#excel_file = "D:\Users\Raphael\.qgis2\python\plugins\PlanningTool\data\excel_data.xlsm"
excel_file = "/Users/Raphael/.qgis2/python/plugins/PlanningTool/data/excel_data.xlsm"

# change cell P21 from 0 to 1 (and vice versa), and the value in E2 should change
# P21 = 1 -> E2 = 0.0031
# P21 = 0 -> E2 = 0.0021

# # openpyxl, save new data to file
# input = 'INPUT - Infra Projects'
# srcfile = openpyxl.load_workbook(excel_file, read_only=False, keep_vba=True)  # to open the excel sheet and if it has macros
# sheet_input = srcfile.get_sheet_by_name(input)  # get sheetname from the file
#
# sheet_input['P21'] = 1
# # this saves and closes the excel file
# srcfile.save(excel_file)


# # run Excel with xlwings
# # this does not work on windows yet, but the only reason is probably that it cannot access the sheet because Excel opens
# # with a put in Product Key Window. So maybe it is actually better to use win32com.client solution as this seems to work better here

#app = xw.App(visible=True) # not necessary
book = xw.Book(excel_file)
input = 'INPUT - Infra Projects'
output = 'Indicator 1 Accessibility'

# set new value
book.sheets[input].range('P21').value = 1
# get new value based on UDF (user defined function)
val = book.sheets[output].range('E2').value
# print value
print val

book.close()  # Ya puedo cerrar el libro.

# # run Excel with win32com.client,
# # this needs to be used in combination with, e.g. openpyxl for saving and getting values
# xl=win32com.client.Dispatch("Excel.Application")
# xl.Workbooks.Open(Filename=excel_file,ReadOnly=0)
# #xl.Application.Run("macrohere")
# xl.Workbooks(1).Close(SaveChanges=1)
# xl.Application.Quit()
# xl=0

# # openpyxl, get value
# with open(excel_file, "rb") as f:
#     in_mem_file = io.BytesIO(f.read())
#
# srcfile = openpyxl.load_workbook(in_mem_file, read_only=True, keep_vba=True, data_only=True)  # to open the excel sheet and if it has macros
# srcfile = openpyxl.load_workbook(excel_file, read_only=True, keep_vba=True, data_only=True)  # to open the excel sheet and if it has macros
# output = 'Indicator 1 Accessibility'
# sheet_output = srcfile.get_sheet_by_name(output)  # get sheetname from the file
#
# val = sheet_output['E2'].value
# print val
#
# srcfile.save(excel_file)


a=5